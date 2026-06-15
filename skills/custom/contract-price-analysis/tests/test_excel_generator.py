"""Tests for the Excel generator (6 sheets + charts)."""

import os

import openpyxl

from scripts.excel_generator import generate_excel


def _sample_groups():
    return [
        {
            "name": "高压开关柜",
            "category": "设备",
            "stats": {
                "count": 3, "mean": 120000.0, "min": 100000.0, "max": 150000.0,
                "median": 120000.0, "std": 20000.0, "outlier_count": 0,
                "outlier_threshold": 165000.0,
            },
            "items": [
                {"goods_name": "高压开关柜", "spec_model": "KYN28", "tech_params": {"电压": "10kV"},
                 "unit_price": 100000.0, "source_contract_no": "C001", "sign_date": "2024-01-01",
                 "supplier": "供应商A"},
                {"goods_name": "10kV高压开关柜", "spec_model": "KYN28", "tech_params": {"电压": "10kV"},
                 "unit_price": 150000.0, "source_contract_no": "C002", "sign_date": "2024-02-01",
                 "supplier": "供应商B", "is_outlier": True},
            ],
        }
    ]


def test_generate_excel_creates_workbook(tmp_path):
    out = tmp_path / "report.xlsx"
    path = generate_excel(_sample_groups(), str(out))
    assert os.path.exists(path)
    assert path.endswith(".xlsx")


def test_generate_excel_has_six_sheets(tmp_path):
    out = tmp_path / "r.xlsx"
    path = generate_excel(_sample_groups(), str(out))
    wb = openpyxl.load_workbook(path)
    assert len(wb.sheetnames) == 6
    expected = {"汇总总表", "分项明细", "图表-价格分布", "图表-价格趋势", "图表-供应商对比", "图表-区间分布"}
    assert set(wb.sheetnames) == expected


def test_summary_sheet_has_header_and_data(tmp_path):
    out = tmp_path / "r.xlsx"
    path = generate_excel(_sample_groups(), str(out))
    wb = openpyxl.load_workbook(path)
    ws = wb["汇总总表"]
    assert ws.cell(1, 1).value == "类别"
    assert ws.cell(2, 2).value == "高压开关柜"
    assert float(ws.cell(2, 4).value) == 120000.0  # mean


def test_items_sheet_marks_outlier(tmp_path):
    out = tmp_path / "r.xlsx"
    path = generate_excel(_sample_groups(), str(out))
    wb = openpyxl.load_workbook(path)
    ws = wb["分项明细"]
    # Row 3 (second item) is the outlier; column 8 = 是否异常
    assert ws.cell(3, 8).value == "是"


def test_empty_groups_does_not_crash(tmp_path):
    out = tmp_path / "empty.xlsx"
    path = generate_excel([], str(out))
    assert os.path.exists(path)
    wb = openpyxl.load_workbook(path)
    assert len(wb.sheetnames) == 6
